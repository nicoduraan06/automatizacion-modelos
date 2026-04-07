from __future__ import annotations

from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook

from src.models.contracts import Excel303Data, TraceCell


SHEET_HINTS = ["iva", "devengado", "deducible", "bases", "resumen", "modelo 303", "303"]

LABEL_HINTS = [
    "base imponible",
    "cuota",
    "devengado",
    "deducible",
    "resultado",
    "compensar",
    "rectificativa",
    "sin actividad",
    "nif",
    "periodo",
    "ejercicio",
]


def _normalize(value: Any) -> str:
    return str(value).strip().lower() if value is not None else ""


class ExcelParser:

    def __init__(self, file_path: Optional[str] = None):
        """
        Permite dos modos:
        - Local: pasar file_path
        - Serverless: usar parse_bytes()
        """
        self.file_path = file_path

    # 🔹 MODO SERVERLESS (Vercel)
    def parse_bytes(self, filename: str, content: bytes) -> Excel303Data:
        wb_values = load_workbook(BytesIO(content), data_only=True)
        wb_formulas = load_workbook(BytesIO(content), data_only=False)

        return self._process_workbooks(filename, wb_values, wb_formulas)

    # 🔹 MODO LOCAL (debug)
    def parse_file(self) -> Excel303Data:
        if not self.file_path:
            raise ValueError("file_path no definido")

        wb_values = load_workbook(self.file_path, data_only=True)
        wb_formulas = load_workbook(self.file_path, data_only=False)

        return self._process_workbooks(self.file_path, wb_values, wb_formulas)

    # 🔹 LÓGICA COMÚN
    def _process_workbooks(self, filename, wb_values, wb_formulas) -> Excel303Data:

        sheets_used: list[str] = []
        cell_index: dict[str, dict[str, Any]] = {}
        label_index: dict[str, list[TraceCell]] = {}

        for ws_values, ws_formula in zip(wb_values.worksheets, wb_formulas.worksheets):
            sheet_name = ws_values.title

            likely_relevant = any(h in _normalize(sheet_name) for h in SHEET_HINTS)
            populated = 0

            for row in ws_values.iter_rows():
                for cell in row:

                    if cell.value is None:
                        continue

                    populated += 1
                    coord = cell.coordinate

                    formula = ws_formula[coord].value
                    if not (isinstance(formula, str) and formula.startswith("=")):
                        formula = None

                    cell_index.setdefault(sheet_name, {})[coord] = {
                        "value": cell.value,
                        "formula": formula,
                    }

                    text = _normalize(cell.value)

                    if any(h in text for h in LABEL_HINTS):
                        trace = TraceCell(
                            sheet=sheet_name,
                            cell=coord,
                            value=cell.value,
                            formula=formula,
                            label=str(cell.value)
                        )

                        label_index.setdefault(text, []).append(trace)
                        likely_relevant = True

            if likely_relevant or populated > 0:
                sheets_used.append(sheet_name)

        return Excel303Data(
            workbook_name=filename,
            sheets_used=sheets_used,
            cell_index=cell_index,
            label_index=label_index,
        )